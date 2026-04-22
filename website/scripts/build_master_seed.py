from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


GSTIN_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]\b")
IGNORE_PRODUCT_WORDS = {"total", "tax", "cgst", "sgst", "igst", "grand total", "amount in words"}
SKIP_ADDRESS_PATTERNS = (
    "si.no",
    "invoice no",
    "invoice date",
    "vehicle no",
    "payment",
    "state code",
    "description",
    "hsn",
    "qty",
    "rate",
    "amount",
)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def maybe_hsn(value: object) -> str:
    text = normalize_text(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 4:
        return digits[:8]
    return ""


def to_float(value: object) -> float:
    text = normalize_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def first_non_empty(cells: Iterable[object]) -> str:
    for cell in cells:
        text = normalize_text(cell)
        if text:
            return text
    return ""


def find_item_header(rows: list[tuple[str, ...]]) -> tuple[int, dict[str, int]] | None:
    for idx, row in enumerate(rows[:45]):
        labels = [normalize_name(cell) for cell in row]
        desc_idx = next((i for i, v in enumerate(labels) if "description" in v or "item" in v), -1)
        hsn_idx = next((i for i, v in enumerate(labels) if "hsn" in v), -1)
        qty_idx = next((i for i, v in enumerate(labels) if v in {"qty", "quantity"}), -1)
        rate_idx = next((i for i, v in enumerate(labels) if "rate" in v or "unit price" in v or "price" in v), -1)
        amount_idx = next((i for i, v in enumerate(labels) if v == "amount"), -1)
        if desc_idx >= 0 and hsn_idx >= 0 and (qty_idx >= 0 or rate_idx >= 0 or amount_idx >= 0):
            return idx, {
                "description": desc_idx,
                "hsn": hsn_idx,
                "qty": qty_idx,
                "rate": rate_idx,
                "amount": amount_idx,
            }
    return None


def extract_products_from_sheet(ws) -> list[dict]:
    raw_rows = [tuple(normalize_text(c) for c in row) for row in ws.iter_rows(values_only=True, max_row=220)]
    if not raw_rows:
        return []
    header_info = find_item_header(raw_rows)
    if not header_info:
        return []
    header_row, col = header_info
    products: list[dict] = []

    for row in raw_rows[header_row + 1 :]:
        if not any(row):
            continue
        desc = row[col["description"]] if col["description"] < len(row) else ""
        hsn = maybe_hsn(row[col["hsn"]] if col["hsn"] < len(row) else "")
        if not desc:
            continue

        desc_lower = desc.lower()
        if any(word in desc_lower for word in IGNORE_PRODUCT_WORDS):
            break
        if len(desc) < 2 or desc_lower in {"description", "item"}:
            continue

        qty = to_float(row[col["qty"]] if col["qty"] >= 0 and col["qty"] < len(row) else 0)
        rate = to_float(row[col["rate"]] if col["rate"] >= 0 and col["rate"] < len(row) else 0)
        amount = to_float(row[col["amount"]] if col["amount"] >= 0 and col["amount"] < len(row) else 0)
        if rate <= 0 and qty > 0 and amount > 0:
            rate = round(amount / qty, 2)

        products.append(
            {
                "description": desc,
                "hsn_code": hsn,
                "price": round(rate, 2) if rate > 0 else 0.0,
            }
        )
    return products


def extract_customers_from_sheet(ws) -> list[dict]:
    rows = [tuple(normalize_text(c) for c in row) for row in ws.iter_rows(values_only=True, max_row=90)]
    if not rows:
        return []

    customer_name = ""
    customer_address = ""
    customer_gstin = ""

    for i, row in enumerate(rows):
        if not any(row):
            continue
        row_joined = " ".join([c for c in row if c])
        gstin_match = GSTIN_RE.search(row_joined)
        if gstin_match:
            found = gstin_match.group(0)
            if "33CHZPS6333N1ZC" not in found:
                customer_gstin = found

        labels = [normalize_name(c) for c in row]
        is_to_row = any(v == "to" for v in labels[:2])
        if not is_to_row:
            continue

        name_candidate = first_non_empty(row[1:]) or first_non_empty(rows[i + 1] if i + 1 < len(rows) else ())
        if not name_candidate or len(name_candidate) < 2:
            continue
        customer_name = name_candidate

        address_parts = []
        for j in range(i + 1, min(i + 8, len(rows))):
            r = rows[j]
            line = first_non_empty(r)
            lower_line = line.lower()
            joined_lower = " ".join([x for x in r if x]).lower()
            if "gstin" in joined_lower or "invoice no" in joined_lower or "invoice date" in joined_lower:
                continue
            if (
                line
                and lower_line not in {"to", "address"}
                and len(line) > 3
                and not any(token in lower_line for token in SKIP_ADDRESS_PATTERNS)
            ):
                address_parts.append(line)
        customer_address = ", ".join(dict.fromkeys(address_parts))[:250]
        break

    return [{"name": customer_name, "address": customer_address, "gstin": customer_gstin}] if customer_name else []


def merge_unique(entries: list[dict], key: str) -> list[dict]:
    merged = {}
    for item in entries:
        value = normalize_text(item.get(key, ""))
        if not value:
            continue
        dedupe_key = normalize_name(value)
        if not dedupe_key:
            continue
        if dedupe_key not in merged:
            merged[dedupe_key] = item
            continue
        current = merged[dedupe_key]
        for field in item:
            if not normalize_text(current.get(field)) and normalize_text(item.get(field)):
                current[field] = item[field]
            if field == "price" and float(item.get("price") or 0) > float(current.get("price") or 0):
                current[field] = item[field]
    return sorted(merged.values(), key=lambda item: normalize_name(str(item.get(key, ""))))


def main() -> int:
    if len(sys.argv) < 4:
        print("Usage: python build_master_seed.py <excel1> <excel2> <output_json>")
        return 1

    file_paths = [Path(sys.argv[1]), Path(sys.argv[2])]
    output_path = Path(sys.argv[3])
    for path in file_paths:
        if not path.exists():
            print(f"Missing file: {path}")
            return 1

    all_products = []
    all_customers = []
    for file_path in file_paths:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            all_products.extend(extract_products_from_sheet(ws))
            all_customers.extend(extract_customers_from_sheet(ws))
        wb.close()

    products = merge_unique(all_products, "description")
    customers = merge_unique(all_customers, "name")
    products = [p for p in products if len(normalize_text(p.get("description", ""))) >= 2]
    customers = [c for c in customers if len(normalize_text(c.get("name", ""))) >= 2]

    payload = {
        "meta": {
            "source_files": [str(p) for p in file_paths],
            "product_count": len(products),
            "customer_count": len(customers),
        },
        "products": products,
        "customers": customers,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote seed: {output_path}")
    print(f"Products: {len(products)}, Customers: {len(customers)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
